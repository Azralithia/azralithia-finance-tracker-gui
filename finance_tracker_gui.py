import os
import sys
import csv
import json
import sqlite3
import logging
from logging.handlers import RotatingFileHandler
from collections import defaultdict
from datetime import datetime, timedelta
from PyQt6.QtGui import (QPalette, QIcon)
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QPushButton,
    QVBoxLayout, QHBoxLayout, QStackedWidget,
    QCheckBox, QLabel, QLineEdit, QComboBox, QDateEdit, 
    QListWidget, QInputDialog, QDialog, QGroupBox,
    QFrame, QTableWidget, QTableWidgetItem, QMessageBox, QFormLayout, 
    QDialogButtonBox, QHeaderView, QRadioButton, QFileDialog, QButtonGroup
)
from PyQt6.QtCore import (
    Qt, QPropertyAnimation, QEasingCurve, QTimer,
    pyqtSignal, pyqtProperty, QSettings, QRect, QDate
)
try:
    import matplotlib.dates as mdates
    from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
    from matplotlib.figure import Figure
    MATPLOTLIB_AVAILABLE = True
except Exception:
    MATPLOTLIB_AVAILABLE = False

logger = logging.getLogger()
logger.setLevel(logging.INFO)
formatter = logging.Formatter(
    "[%(levelname)s] %(asctime)s - Call Line: %(lineno)d - %(message)s",
    "%H:%M:%S"
)

console_handler = logging.StreamHandler()
console_handler.setFormatter(formatter)
logger.addHandler(console_handler)

file_handler = RotatingFileHandler(
    "finance_tracker.log",
    maxBytes=1_000_000,
    backupCount=5
)
file_handler.setFormatter(formatter)
logger.addHandler(file_handler)

try:
    import openpyxl  
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl not found. Install with `pip install openpyxl` to enable Excel export.")


# ---------------------------
#           Config
# ---------------------------
BUTTON_WIDTH = 48
BUTTON_HEIGHT = 40
SIDEBAR_COLLAPSED_WIDTH = 60
SIDEBAR_EXPANDED_WIDTH = 300
ANIM_DURATION = 250
TOGGLE_ANIM_DURATION = 200
FADE_DURATION = 200
TOGGLE_WIDTH = 60
TOGGLE_HEIGHT = 28
TOGGLE_MARGIN = 2
TOGGLE_PADDING = 6
DARK_MODE = """
            .settings-label {
                font-size: 16px;
                color: white; 
                padding: 8px;
                border: 1px solid #555; 
                border-right: none;
                border-bottom: none;
            }
            .settings-value-widget {
                padding: 4px;
                border: 1px solid #555; 
                border-bottom: none;
            }
            .settings-row-frame {
                border: none;
            }
            .settings-row-frame:last-child .settings-label,
            .settings-row-frame:last-child .settings-value-widget {
                border-bottom: 1px solid #555; 
            }
            #SummaryCard {
                border: 1px solid #3a3a3a;
                border-radius: 10px;
                background: #2b2b2b;
            }
            #SummaryCardTitle { 
                font-size: 13px; 
                color: #cfcfcf; 
            }
            #SummaryCardValue { 
                font-size: 24px; 
                font-weight: 600; 
                color: #ffffff; 
            }
            #SummaryGroup { 
                border: 1px solid #3a3a3a; 
                border-radius: 10px; 
            }
            #SummaryLabelSmall { 
            color: #e6e6e6; 
            }
            #SummaryRow { 
                color: #eaeaea;
                font-size: 16px;
                font-weight: 500;
                padding: 4px 0px;
            }
            QWidget { 
                background-color: #2c2c2c; 
                color: white; 
            }
            
            QWidget#SidebarBottom { 
                background-color: #1a1a1a; 
            }     
            
            QWidget#Sidebar { 
                background-color: #1a1a1a; 
                
            }

            #Sidebar QPushButton {
                background-color: transparent;
            }
            #Sidebar QPushButton:hover {
                background-color: rgba(255,255,255,0.08);
            }
            #Sidebar QPushButton:pressed {
                background-color: rgba(255,255,255,0.16);
            }        
            
            /* Sidebar buttons */
            QPushButton { 
                color: white; 
                background-color: #2c2c2c;
                border: none; 
                border-radius: 4px; 
                text-align: left;
                font-size: 14px;
                padding: 6px;
                padding-left: 10px; 
            }
            /* Submenu indentation */
            QPushButton[subitem="true"] {
                padding-left: 30px;
            }
            QPushButton:hover { 
                background-color: #404040; 
            }
            QPushButton:pressed { 
                background-color: #606060;    
            }
            QLabel, QCheckBox { 
                color: white; 
            }
            QLineEdit { 
                background-color: #3a3a3a;
                border: 1px solid #555;
                border-radius: 4px;
                padding: 4px;
                color: white;
            }
            QTableWidget {
                gridline-color: #555;
                background-color: #2c2c2c;
                alternate-background-color: #3a3a3a;
                color: white;
                selection-background-color: #505050;
            }
            QHeaderView::section {
                background-color: #3a3a3a;
                color: white;
                padding: 4px;
                border: none;
            }
            QComboBox {
                background-color: #3a3a3a;
                border: 1px solid #555;
                border-radius: 4px;
                padding: 4px;
                color: white;
            }
            QComboBox QAbstractItemView {
                background-color: #2c2c2c;
                selection-background-color: #404040;
                color: white;
            }
        """
LIGHT_MODE = """
            .settings-label {
                font-size: 16px;
                color: black;
                padding: 8px;
                border: 1px solid #ccc; 
                border-right: none;
                border-bottom: none;
            }
            .settings-value-widget {
                padding: 4px;
                border: 1px solid #ccc; 
                border-bottom: none;
            }
            .settings-row-frame {
                border: none;
            }
            .settings-row-frame:last-child .settings-label,
            .settings-row-frame:last-child .settings-value-widget {
                border-bottom: 1px solid #ccc;
            }
            #SummaryCard {
                border: 1px solid #d0d0d0;
                border-radius: 10px;
                background: #ffffff;
            }
            #SummaryCardTitle { 
                font-size: 13px; 
                color: #333; 
            }
            #SummaryCardValue { 
                font-size: 24px; 
                font-weight: 600; 
                color: #111; 
            }
            #SummaryGroup { 
                border: 1px solid #e0e0e0; 
                border-radius: 10px; 
            }
            #SummaryLabelSmall { 
            color: #222; 
            }
            #SummaryRow { 
                color: #222;
                font-size: 16px;
                font-weight: 500;
                padding: 4px 0px;
            }
            QWidget { 
                background-color: #f0f0f0; 
                color: black; 
            }
            
            QWidget#SidebarBottom { 
                background-color: #e6e6e6; 
            }
            
            QWidget#Sidebar { 
                background-color: #e6e6e6; 
            }

            #Sidebar QPushButton {
                background-color: transparent;
            }
            #Sidebar QPushButton:hover {
                background-color: rgba(0,0,0,0.06);
            }
            #Sidebar QPushButton:pressed {
                background-color: rgba(0,0,0,0.12);
            }

            /* Sidebar buttons */
            QPushButton { 
                color: black;                      
                background-color: #f5f5f5; 
                border: none; 
                border-radius: 4px;
                text-align: left;
                font-size: 14px;
                padding: 6px;
                padding-left: 10px;               
            }
            /* Submenu indentation */
            QPushButton[subitem="true"] {
                padding-left: 30px;
            }
            QPushButton:hover { 
                background-color: #dcdcdc; 
            }
            QPushButton:pressed { 
                background-color: #c0c0c0;    
            }
            QLabel, QCheckBox { 
                color: black; 
            }
            QLineEdit { 
                background-color: white;
                border: 1px solid #aaa;
                border-radius: 4px;
                padding: 4px;
                color: black;
            }
            QTableWidget {
                gridline-color: #ccc;
                background-color: white;
                alternate-background-color: #f5f5f5;
                color: black;
                selection-background-color: #d0d0d0;
            }
            QHeaderView::section {
                background-color: #e0e0e0;
                color: black;
                padding: 4px;
                border: none;
            }
            QComboBox {
                background-color: white;
                border: 1px solid #aaa;
                border-radius: 4px;
                padding: 4px;
                color: black;
            }
            QComboBox QAbstractItemView {
                background-color: #fff;
                selection-background-color: #dcdcdc;
                color: black;
            }
        """

def create_db():
    conn = sqlite3.connect("transactions.db")
    cursor = conn.cursor()
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS transactions (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        type TEXT NOT NULL,
        amount REAL NOT NULL,
        category TEXT,
        date DATETIME NOT NULL,
        note TEXT
    )
    """)
    conn.commit()
    conn.close()

def map_display_format(display_format: str) -> str:
    if display_format == "(DD-MM-YYYY) | Day-Month-Year":
        return "dd-MM-yyyy"
    if display_format == "(MM-DD-YYYY) | Month-Day-Year":
        return "MM-dd-yyyy"
    if display_format == "(YYYY-DD-MM) | Year-Day-Month":
        return "yyyy-dd-MM"
    return "yyyy-MM-dd" # Default to Year-Month-Day

def graph_format(display_format: str) -> str:
    if display_format == "dd-MM-yyyy":
        return "%d-%m-%Y"
    if display_format == "MM-dd-yyyy":
        return "%m-%d-%Y"
    if display_format == "yyyy-dd-mm":
        return "%Y-%d-%m"
    return "%Y-%m-%d" # Default to Year-Month-Day

def format_date_for_display(date_str: str) -> str:
    s = QSettings("Azralithia", "FinanceTracker").value("date_format", "(YYYY-MM-DD) | Year-Month-Day")
    
    if " | " in s:
        s = s.split(" | ")[0].strip("()") 
    try:
        y, m, d = date_str.split("-")
    except Exception:
        return date_str
    
    if s == "DD-MM-YYYY":
        return f"{d}-{m}-{y}"
    if s == "MM-DD-YYYY":
        return f"{m}-{d}-{y}"
    if s == "YYYY-DD-MM":
        return f"{y}-{d}-{m}"
    return f"{y}-{m}-{d}" # Default to Year-Month-Day

def resource_path(relative_path: str) -> str:
    try:
        base_path = sys._MEIPASS
    except AttributeError:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

# --- Pages ---
class MainPage(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.db_path = "transactions.db"
        self._stack = None
        self._build_ui()
        self.load_summary()

    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(20)

        self.welcome_label = QLabel("👋 Welcome to Azralithia Finance Tracker!")
        self.welcome_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.welcome_label.setStyleSheet("font-size: 24px; font-weight: bold;")
        layout.addWidget(self.welcome_label)

        # Summary group boxes
        self.gb_last7 = self._make_group_box("Last 7 Days Summary")
        self.gb_month = self._make_group_box("Current Month Balance")

        summary_layout = QHBoxLayout()
        summary_layout.addWidget(self.gb_last7)
        summary_layout.addWidget(self.gb_month)
        layout.addLayout(summary_layout)

        # Quick add button
        self.add_btn = QPushButton("➕ Add New Transaction")
        self.add_btn.setFixedHeight(50)
        self.add_btn.setStyleSheet("""
            QPushButton {
                background-color: #1976D2; /* Blue */
                color: white;
                border: none;
                border-radius: 8px;
                font-size: 14px;
                font-weight: bold;
                text-align: center;
                padding: 10px 20px;
            }
            QPushButton:hover {
                background-color: #2196F3;
            }
            QPushButton:pressed {
                background-color: #0D47A1;
            }
        """)
        layout.addWidget(self.add_btn, alignment=Qt.AlignmentFlag.AlignCenter)

        self.add_btn.clicked.connect(self._open_manage_transactions)

    def _make_group_box(self, title):
        gb = QGroupBox(title)
        gb.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                font-size: 16px;
                border: 1px solid #555;
                border-radius: 8px;
                margin-top: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                subcontrol-position: top center;
                padding: 0 3px;
            }
        """)
        gb_layout = QVBoxLayout(gb)
        gb_layout.setContentsMargins(15, 30, 15, 15)
        gb_layout.setSpacing(10)

        gb._income_label = QLabel("Income: $0.00")
        gb._expense_label = QLabel("Expense: $0.00")
        gb._balance_label = QLabel("Balance: $0.00")

        for lbl in (gb._income_label, gb._expense_label, gb._balance_label):
            lbl.setStyleSheet("font-size: 14px;")
            gb_layout.addWidget(lbl)

        return gb

    def load_summary(self):
        conn = sqlite3.connect(self.db_path)
        cur = conn.cursor()

        # Last 7 days
        last7_start = QDate.currentDate().addDays(-6).toString("yyyy-MM-dd")
        last7_end = QDate.currentDate().toString("yyyy-MM-dd")
        cur.execute("""
            SELECT type, SUM(amount) FROM transactions
            WHERE date BETWEEN ? AND ?
            GROUP BY type
        """, (last7_start, last7_end))
        last7_data = dict(cur.fetchall())
        income_7 = last7_data.get("income", 0)
        expense_7 = last7_data.get("expense", 0)
        balance_7 = income_7 - expense_7
        self.gb_last7._income_label.setText(f"Income: ${income_7:.2f}")
        self.gb_last7._expense_label.setText(f"Expense: ${expense_7:.2f}")
        self.gb_last7._balance_label.setText(f"Balance: ${balance_7:.2f}")

        today = QDate.currentDate()
        month_start = today.toString("yyyy-MM-01")
        month_end = today.toString("yyyy-MM-dd")
        cur.execute("""
            SELECT type, SUM(amount) FROM transactions
            WHERE date BETWEEN ? AND ?
            GROUP BY type
        """, (month_start, month_end))
        month_data = dict(cur.fetchall())

        income_m = month_data.get("income", 0) or 0
        expense_m = month_data.get("expense", 0) or 0
        balance_m = income_m - expense_m

        self.gb_month._income_label.setText(f"Income: ${income_m:.2f}")
        self.gb_month._expense_label.setText(f"Expense: ${expense_m:.2f}")
        self.gb_month._balance_label.setText(f"Balance: ${balance_m:.2f}")

        conn.close()

    def set_page_switcher(self, stack_widget, transactions_page_widget):
        self._stack = stack_widget
        self._transactions_page = transactions_page_widget

    def _open_manage_transactions(self):
        self._stack.setCurrentWidget(self._transactions_page)

class TransactionsPage(QWidget):
    data_changed = pyqtSignal()
    def __init__(self, main_window = None):
        super().__init__()
        self.main_window = main_window
        self.settings = QSettings("Azralithia", "FinanceTracker")
        layout = QVBoxLayout(self)

        # Income/Expense selector
        self.type_row = QHBoxLayout()
        self.income_btn = QPushButton("Income")
        self.expense_btn = QPushButton("Expense")

        self.type_button_group = QButtonGroup(self)
        self.type_button_group.setExclusive(True)
        self.type_button_group.addButton(self.income_btn)
        self.type_button_group.addButton(self.expense_btn)

        for btn in (self.income_btn, self.expense_btn):
            btn.setCheckable(True)
            btn.setFixedHeight(35)
            self.type_row.addWidget(btn)
            btn.setMaximumWidth(250)
        layout.addLayout(self.type_row)
        self.type_button_group.buttonClicked.connect(lambda btn: self.set_type(btn.text()))

        # Categories
        self.categories = {}
        self.load_categories()

        # Default type = Expense
        self.current_type = "Expense"
        self.expense_btn.setChecked(True)
        self.update_type_styles()

        # Category dropdown + edit button
        cat_row = QHBoxLayout()
        self.category = QComboBox()
        self.edit_cat_btn = QPushButton(" ✏️")
        self.edit_cat_btn.setFixedWidth(BUTTON_WIDTH)
        cat_row.addWidget(self.category)
        cat_row.addWidget(self.edit_cat_btn)
        layout.addLayout(cat_row)
        self.reload_categories()

        # Form fields
        self.amount = QLineEdit()
        self.amount.setPlaceholderText("Amount")
        self.date = QDateEdit()
        self.date.setCalendarPopup(True)
        self.date.setDate(QDate.currentDate())

        settings = QSettings("Azralithia", "FinanceTracker")
        current_display_format = settings.value("date_format", "(YYYY-MM-DD) | Year-Month-Day")
        display_format = map_display_format(current_display_format)
        self.date.setDisplayFormat(display_format)

        self.notes = QLineEdit()
        self.notes.setPlaceholderText("Notes (optional)")

        for widget in (self.amount, self.date, self.notes):
            layout.addWidget(widget)

        save_btn_layout = QHBoxLayout()
        self.save_btn = QPushButton("💾 Save Transaction")
        self.save_btn.setFixedHeight(35) 
        self.save_btn.setStyleSheet("""
            QPushButton {
                background-color: #1976D2; /* Blue */
                color: white;
                border: none;
                border-radius: 8px;
                font-size: 14px;
                font-weight: bold;
                text-align: center;
                padding: 10px 20px;
            }
            QPushButton:hover {
                background-color: #2196F3;
            }
            QPushButton:pressed {
                background-color: #0D47A1;
            }
        """)
        save_btn_layout.addStretch()
        save_btn_layout.addWidget(self.save_btn)
        save_btn_layout.addStretch()
        layout.addLayout(save_btn_layout)

        self.feedback = QLabel("")
        layout.addWidget(self.feedback)

        # Hookups
        self.edit_cat_btn.clicked.connect(self.edit_categories)
        self.save_btn.clicked.connect(self.save_transaction)

        # Recent transactions preview
        recent_label = QLabel("Recent transactions")
        recent_label.setStyleSheet("font-weight:600; margin-top:8px;")
        layout.addWidget(recent_label)

        self.recent_table = QTableWidget()
        self.recent_table.setColumnCount(6)
        self.recent_table.setHorizontalHeaderLabels(["Date", "Type", "Category", "Amount", "Note", "Actions"])
        self.recent_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        header = self.recent_table.horizontalHeader()
        for i in range(5):
            header.setSectionResizeMode(i, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(5, QHeaderView.ResizeMode.Fixed)
        self.recent_table.setColumnWidth(5, 100)
        self.recent_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        layout.addWidget(self.recent_table)
        self.recent_table.cellClicked.connect(self._handle_recent_table_cell_click)
    
    def _handle_recent_table_cell_click(self, row, column):
            transaction_id_item = self.recent_table.item(row, 0)
            transaction_id = transaction_id_item.data(Qt.ItemDataRole.UserRole)
            conn = sqlite3.connect("transactions.db")
            cur = conn.cursor()
            cur.execute("SELECT date, type, category, amount, note FROM transactions WHERE id=?", (transaction_id,))
            data = cur.fetchone()
            conn.close()
            if not data:
                return
            date, ttype, category, amount, note = data
            dialog = TransactionEditDialog(transaction_id, "transactions.db", self)

            if column == 0: 
                dialog.date_edit.setDate(QDate.fromString(date, "yyyy-MM-dd"))
                dialog.date_edit.setFocus()
            elif column == 1: 
                dialog.type_combo.setCurrentText(ttype.title())
                dialog.type_combo.setFocus()
            elif column == 2: 
                dialog.category_combo.setCurrentText(category.title())
                dialog.category_combo.setFocus()
            elif column == 3: 
                dialog.amount_edit.setText(str(amount))
                dialog.amount_edit.setFocus()
            elif column == 4: 
                dialog.note_edit.setText(note)
                dialog.note_edit.setFocus()

            if dialog.exec():
                self.load_recent_transactions() 
                self.data_changed.emit() 
    
    def showEvent(self, event):
            self.load_recent_transactions()
            super().showEvent(event)

    def load_recent_transactions(self, limit=10):
        conn = sqlite3.connect("transactions.db")
        cur = conn.cursor()
        cur.execute("SELECT id, date, type, category, amount, COALESCE(note,'') FROM transactions ORDER BY id DESC LIMIT ?", (limit,))
        rows = cur.fetchall()
        conn.close()
        self.recent_table.setRowCount(len(rows))
        for r, (rid, date, t, cat, amt, note) in enumerate(rows):
            date_item = QTableWidgetItem(format_date_for_display(date))
            date_item.setData(Qt.ItemDataRole.UserRole, rid)
            self.recent_table.setItem(r, 0, date_item)
            self.recent_table.setItem(r, 1, QTableWidgetItem(t.title()))
            self.recent_table.setItem(r, 2, QTableWidgetItem((cat or "").title()))
            self.recent_table.setItem(r, 3, QTableWidgetItem(f"{amt:.2f}"))
            self.recent_table.setItem(r, 4, QTableWidgetItem(note))
            btns_widget = QWidget()
            btns_layout = QHBoxLayout(btns_widget)
            btns_layout.setContentsMargins(0, 0, 0, 0)
            btns_layout.setSpacing(6)
            if self.main_window and rid in self.main_window.pending_delete_transactions:
                entry = self.main_window.pending_delete_transactions[rid]
                labels_map = entry.setdefault("countdown_labels", {})

                if "recent" in labels_map and labels_map["recent"] is not None:
                    countdown_label = labels_map["recent"]
                else:
                    countdown_label = QLabel(str(entry["countdown"]))
                    countdown_label.setFixedWidth(20)
                    countdown_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
                    countdown_label.setStyleSheet("font-weight: bold; color: red;")
                    labels_map["recent"] = countdown_label

                buttons_map = entry.setdefault("undo_buttons", {})
                undo_btn = QPushButton("Undo")
                undo_btn.setCursor(Qt.CursorShape.PointingHandCursor)
                undo_btn.setToolTip("Undo Delete")
                undo_btn.clicked.connect(lambda _, _rid=rid: self.main_window.undo_delete(_rid))
                buttons_map["recent"] = undo_btn
                inline_undo_widget = QWidget()
                inline_undo_layout = QHBoxLayout(inline_undo_widget)
                inline_undo_layout.setContentsMargins(0, 0, 0, 0)
                inline_undo_layout.setSpacing(2)
                inline_undo_layout.addWidget(undo_btn)
                inline_undo_layout.addWidget(countdown_label)
                inline_undo_layout.addStretch()
                btns_layout.addWidget(inline_undo_widget)
            else:
                edit_btn = QPushButton(" ✏️")
                delete_btn = QPushButton(" 🗑️")
                edit_btn.setFixedSize(BUTTON_WIDTH, BUTTON_HEIGHT - 10)
                delete_btn.setFixedSize(BUTTON_WIDTH, BUTTON_HEIGHT - 10)
                edit_btn.setCursor(Qt.CursorShape.PointingHandCursor)
                delete_btn.setCursor(Qt.CursorShape.PointingHandCursor)
                edit_btn.setToolTip("Edit")
                delete_btn.setToolTip("Delete")
                edit_btn.clicked.connect(lambda _, _rid=rid: self._edit_from_preview(_rid))
                delete_btn.clicked.connect(lambda _, _rid=rid, _date=date, _type=t, _cat=cat, _amt=amt, _note=note:
                    self.main_window.mark_transaction_for_deletion( _rid, (_rid, _date, _type, _cat, _amt, _note))
                )
                btns_layout.addWidget(edit_btn)
                btns_layout.addWidget(delete_btn)
            btns_layout.addStretch()
            self.recent_table.setCellWidget(r, 5, btns_widget)

    def _edit_from_preview(self, rid):
        dlg = TransactionEditDialog(rid, "transactions.db", self)
        if dlg.exec():
            self.load_recent_transactions()
            self.data_changed.emit()

    # -=- Helpers -=-
    def set_type(self, t: str):
        self.current_type = t
        self.income_btn.setChecked(t == "Income")
        self.expense_btn.setChecked(t == "Expense")
        self.update_type_styles()
        self.reload_categories()

    def update_type_styles(self):
        self.income_btn.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50; /* Green */
                color: white;
                border: 1px solid #388E3C;
                border-radius: 4px;
                text-align: center; 
                font-weight: bold;
                padding-left: 0px; 
            }
            QPushButton:checked {
                background-color: #66BB6A;
                border: 2px solid #388E3C;
            }
            QPushButton:hover {
                background-color: #81C784;
            }
        """)
        self.expense_btn.setStyleSheet("""
            QPushButton {
                background-color: #F44336; /* Red */
                color: white;
                border: 1px solid #D32F2F;
                border-radius: 4px;
                text-align: center; 
                font-weight: bold;
                padding-left: 0px; 
            }
            QPushButton:checked {
                background-color: #EF5350;
                border: 2px solid #D32F2F;
            }
            QPushButton:hover {
                background-color: #E57373;
            }
        """)

    def reload_categories(self):
        self.category.clear()
        self.category.addItems(self.categories.get(self.current_type, []))

    def edit_categories(self):
        current = self.categories.get(self.current_type, [])
        dlg = CategoryEditor(current, self)
        if dlg.exec():
            updated = dlg.get_categories()
            cleaned = []
            seen = set()
            for name in (s.strip() for s in updated):
                if name and name not in seen:
                    seen.add(name)
                    cleaned.append(name)
            self.categories[self.current_type] = cleaned or ["Other"]
            self.reload_categories()
            self.save_categories()  

    def load_categories(self):
        stored = self.settings.value("categories", None)
        defaults = {
            "Income": ["Salary", "Gift", "Bonus", "Other"],
            "Expense": ["Food", "Rent", "Utilities", "Transport", "Misc"]
        }
        if not stored:
            self.categories = defaults
            return
        try:
            data = json.loads(stored)
            if not isinstance(data, dict):
                raise ValueError
            for key in ("Income", "Expense"):
                if key not in data or not isinstance(data[key], list) or not data[key]:
                    data[key] = defaults[key]
            self.categories = data
        except Exception:
            self.categories = defaults

    def save_categories(self):
        self.settings.setValue("categories", json.dumps(self.categories))

    def save_transaction(self):
        try:
            amount = float(self.amount.text())
            if amount <= 0:
                raise ValueError("Amount must be positive.")
        except ValueError as e:
            self.feedback.setText(f"❌ {e}")
            return

        tx = {
            "type": self.current_type.lower(),
            "amount": amount,
            "category": self.category.currentText().lower(),
            "date": self.date.date().toString("yyyy-MM-dd"),
            "note": self.notes.text().strip()
        }
        conn = sqlite3.connect("transactions.db")
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO transactions (type, amount, category, date, note)
            VALUES (?, ?, ?, ?, ?)
        """, (tx['type'], tx['amount'], tx['category'], tx['date'], tx['note']))
        conn.commit()
        conn.close()
        logging.getLogger().info(f"Transaction saved: {tx}")
        self.feedback.setText("✅ Transaction saved!")
        self.amount.clear()
        self.notes.clear()
        self.data_changed.emit()
        self.load_recent_transactions()
        QTimer.singleShot(5000, lambda: self.feedback.setText(""))
    
    def apply_theme(self, mode: str):
        if mode == "dark":
            self.setStyleSheet(DARK_MODE)
        else:
            self.setStyleSheet(LIGHT_MODE)

class SummaryPage(QWidget):
    def __init__(self, db_path="transactions.db", parent=None):
        super().__init__(parent)
        self.db_path = db_path
        self._build_ui()

        self.start_picker.dateChanged.connect(self.refresh_summary)
        self.end_picker.dateChanged.connect(self.refresh_summary)
        self.refresh_summary()

    def set_db_path(self, db_path: str):
        self.db_path = db_path
        self.refresh_summary()

    def _build_ui(self):
        outer = QVBoxLayout(self)
        outer.setContentsMargins(16, 16, 16, 16)
        outer.setSpacing(16)

        header = QHBoxLayout()
        header.setSpacing(12)

        settings = QSettings("Azralithia", "FinanceTracker")
        light_mode = settings.value("light_mode", False, type=bool)
        current_display_format = settings.value("date_format", "(YYYY-MM-DD) | Year-Month-Day") 
        display_format = map_display_format(current_display_format)

        self.start_label = QLabel("📅 From:")
        self.start_label.setObjectName("SummaryLabelSmall")
        self.start_picker = QDateEdit()
        self.start_picker.setCalendarPopup(True)
        self.start_picker.setDisplayFormat(display_format)
        self.start_picker.setDate(QDate.currentDate().addMonths(-1))

        self.end_label = QLabel("To:")
        self.end_picker = QDateEdit()
        self.end_picker.setCalendarPopup(True)
        self.end_picker.setDisplayFormat(display_format) 
        self.end_picker.setDate(QDate.currentDate())

        header.addWidget(self.start_label)
        header.addWidget(self.start_picker)
        header.addWidget(self.end_label)
        header.addWidget(self.end_picker)

        self.refresh_btn = QPushButton("Refresh")
        self.refresh_btn.clicked.connect(self.refresh_summary)

        # Totals
        totals = QHBoxLayout()
        totals.setSpacing(12)

        self.card_income = self._make_metric_card("Income", "0.00")
        self.card_expense = self._make_metric_card("Expense", "0.00")
        self.card_balance = self._make_metric_card("Balance", "0.00")

        totals.addWidget(self.card_income)
        totals.addWidget(self.card_expense)
        totals.addWidget(self.card_balance)

        self.gb_income = self._make_group_box("Income Breakdown")
        self.gb_expense = self._make_group_box("Expense Breakdown")
        self.gb_income.setObjectName("SummaryCardTitle") 
        self.gb_expense.setObjectName("SummaryCardTitle")
        
        breakdowns_layout = QHBoxLayout()
        breakdowns_layout.addWidget(self.gb_income)
        breakdowns_layout.addWidget(self.gb_expense)

        # Graph 
        if MATPLOTLIB_AVAILABLE:
            self.graph_fig = Figure(figsize=(4,3), tight_layout=True)
            self.graph_canvas = FigureCanvas(self.graph_fig)
            outer.addWidget(self.graph_canvas)
        else:
            self.graph_placeholder = QLabel("Graph unavailable (matplotlib not installed)")
            outer.addWidget(self.graph_placeholder)
        outer.addLayout(header)
        outer.addLayout(totals)
        outer.addWidget(self._hline())
        outer.addLayout(breakdowns_layout)
        self.apply_theme("light" if light_mode else "dark")
    
    def _make_metric_card(self, title: str, value: str) -> QGroupBox:
        box = QGroupBox()
        box.setObjectName("SummaryCard")
        lay = QVBoxLayout(box)
        lay.setContentsMargins(16, 12, 16, 16)
        lay.setSpacing(6)

        lbl_title = QLabel(title)
        lbl_title.setObjectName("SummaryCardTitle")

        lbl_value = QLabel(value)
        lbl_value.setObjectName("SummaryCardValue")

        lay.addWidget(lbl_title)
        lay.addWidget(lbl_value)
        lay.addStretch()

        box._value_label = lbl_value
        return box

    def _make_group_box(self, title: str) -> QGroupBox:
        gb = QGroupBox(title)
        gb.setObjectName("SummaryGroup")
        v = QVBoxLayout(gb)
        v.setContentsMargins(16, 20, 16, 16)  
        v.setSpacing(8)  

        inner = QWidget()
        inner_layout = QVBoxLayout(inner)
        inner_layout.setContentsMargins(0, 0, 0, 0)
        inner_layout.setSpacing(6)  

        v.addWidget(inner)

        gb._rows_container = inner
        gb._rows_layout = inner_layout
        return gb

    def _hline(self):
        line = QFrame()
        line.setFrameShape(QFrame.Shape.HLine)
        line.setFrameShadow(QFrame.Shadow.Sunken)
        return line

    def _plot_balance_over_range(self, start_date, end_date):
        if not MATPLOTLIB_AVAILABLE:
            return
        settings = QSettings("Azralithia", "FinanceTracker")
        light_mode = settings.value("light_mode", False, type=bool)
        conn = self._connect()
        cur = conn.cursor()
        cur.execute("""
            SELECT date,
            SUM(CASE WHEN lower(type)='income' THEN amount WHEN lower(type)='expense' THEN -amount ELSE 0 END) as net
            FROM transactions
            WHERE date >= ? AND date <= ?
            GROUP BY date
            ORDER BY date
        """, (start_date, end_date))
        rows = cur.fetchall()
        conn.close()

        if rows:
            rows_map = {r[0]: r[1] for r in rows}
            d0, d1 = datetime.strptime(start_date, "%Y-%m-%d"), datetime.strptime(end_date, "%Y-%m-%d")
            raw_dates, nets, running = [], [], 0.0
            cur_date = d0
            while cur_date <= d1:
                s = cur_date.strftime("%Y-%m-%d")
                running += float(rows_map.get(s, 0.0))
                raw_dates.append(cur_date)
                nets.append(running)
                cur_date += timedelta(days=1)
            dates = mdates.date2num(raw_dates)
        else:
            dates, nets = [], []

        self.graph_fig.clf()
        ax = self.graph_fig.add_subplot(111)

        palette = self.palette()
        bg_color = palette.color(QPalette.ColorRole.Window).name()
        text_color = palette.color(QPalette.ColorRole.WindowText).name()
        line_color = "#1f77b4" if text_color == "#000000" else "#88c0d0"
        grid_color = "#ccc" if text_color == "#000000" else "#4a4a4a"

        ax.set_facecolor(bg_color)
        self.graph_fig.set_facecolor(bg_color)
        ax.tick_params(axis='x', colors=text_color)
        ax.tick_params(axis='y', colors=text_color)
        ax.yaxis.label.set_color(text_color)
        ax.xaxis.label.set_color(text_color)
        ax.title.set_color(text_color)
        for spine in ax.spines.values():
            spine.set_edgecolor(text_color)
        ax.grid(True, color=grid_color)

        if len(dates) > 0:
            ax.plot(dates, nets, color=line_color)
            ax.set_title("Balance over Time")
            ax.set_xlabel("Date")
            ax.set_ylabel("Balance")
            settings = QSettings("Azralithia", "FinanceTracker")
            current_display_format = settings.value("date_format", "(YYYY-MM-DD) | Year-Month-Day")
            strftime_format = graph_format(map_display_format(current_display_format))
            ax.xaxis.set_major_formatter(mdates.DateFormatter(strftime_format))
            self.graph_fig.autofmt_xdate()
        else:
            ax.text(0.5, 0.5, "No data for selected range", ha='center', va='center',
                    color='white' if not light_mode else 'black')

        self.graph_canvas.draw_idle()

    # --=-- Data / Queries --=--
    def _get_date_range(self):
        sd = self.start_picker.date().toString("yyyy-MM-dd")
        ed = self.end_picker.date().toString("yyyy-MM-dd")
        return sd, ed

    def _connect(self):
        if not self.db_path:
            raise RuntimeError("SummaryPage: db_path not set. Call set_db_path(path_to_sqlite).")
        return sqlite3.connect(self.db_path)

    def _fetch_totals_and_counts(self, start_date: str, end_date: str):
        totals = {'income': 0.0, 'expense': 0.0, 'balance': 0.0}
        counts_income = defaultdict(lambda: {'total': 0.0, 'count': 0})
        counts_expense = defaultdict(lambda: {'total': 0.0, 'count': 0})

        query_totals = """
            SELECT lower(type) AS type, COALESCE(SUM(amount),0) AS total
            FROM transactions
            WHERE date >= ? AND date <= ?
            GROUP BY lower(type)
        """
        query_counts = """
            SELECT lower(type) AS type, lower(category) AS category,
                COALESCE(SUM(amount),0) AS total, COUNT(*) AS count
            FROM transactions
            WHERE date >= ? AND date <= ?
            GROUP BY lower(type), lower(category)
            ORDER BY total DESC
        """
        ym = (start_date, end_date)
        with self._connect() as conn:
            cur = conn.cursor()
            # Overall totals
            for t, total in cur.execute(query_totals, ym):
                if t.lower() == "income":
                    totals['income'] = float(total or 0)
                elif t.lower() == "expense":
                    totals['expense'] = float(total or 0)
            totals['balance'] = totals['income'] - totals['expense']

            # Detailed counts per category
            for t, cat, total_amt, count in cur.execute(query_counts, ym):
                if (t or "").lower() == "income":
                    counts_income[cat or "Uncategorized"] = {
                        'total': float(total_amt or 0),
                        'count': int(count or 0)
                    }
                elif (t or "").lower() == "expense":
                    counts_expense[cat or "Uncategorized"] = {
                        'total': float(total_amt or 0),
                        'count': int(count or 0)
                    }

        return totals, counts_income, counts_expense

    # ----- Refresh / Render -----
    def refresh_summary(self):
        try:
            sd, ed = self._get_date_range()
            totals, c_in, c_ex = self._fetch_totals_and_counts(sd, ed)

            self.card_income._value_label.setText(f"{totals['income']:.2f}")
            self.card_expense._value_label.setText(f"{totals['expense']:.2f}")
            self.card_balance._value_label.setText(f"{totals['balance']:.2f}")

            self._render_breakdown(self.gb_income, c_in)
            self._render_breakdown(self.gb_expense, c_ex)
            self._plot_balance_over_range(sd, ed)

        except Exception as e:
            # Show an error in the UI instead of crashing
            self.card_income._value_label.setText("—")
            self.card_expense._value_label.setText("—")
            self.card_balance._value_label.setText("—")
            self._render_breakdown(self.gb_income, {"Error": {'total': 0.0, 'count': 0}}) 
            self._render_breakdown(self.gb_expense, {f"Error: {str(e)}": {'total': 0.0, 'count': 0}}) 
            logger.error(f"Error refreshing summary: {e}", exc_info=True)

    def _render_breakdown(self, groupbox: QGroupBox, data: dict):
        lay = groupbox._rows_layout
        while (item := lay.takeAt(0)) is not None:
            w = item.widget()
            if w:
                w.deleteLater()

        if not data:
            lbl = QLabel("No data.")
            lbl.setObjectName("SummaryRow")
            lay.addWidget(lbl)
            return

        for cat, info in data.items():
            total = info['total']
            count = info['count']
            capitalized_cat = cat.title() if cat else "Uncategorized"
            row = QLabel(f"{capitalized_cat} ({count}) — {total:.2f}")
            row.setObjectName("SummaryRow")
            lay.addWidget(row)
    
    def apply_theme(self, mode: str):
        if mode == "dark":
            self.setStyleSheet(DARK_MODE)
        else:
            self.setStyleSheet(LIGHT_MODE)
        self.refresh_summary()

class HistoryPage(QWidget):
    data_changed = pyqtSignal()
    def __init__(self, db_path="transactions.db", parent=None, main_window = None):
        super().__init__(parent)
        self.db_path = db_path
        self.main_window = main_window
        self.page_size = 10
        self.current_page = 0
        self.total_rows = 0
        self._build_ui()
        self.load_page()
        settings = QSettings("Azralithia", "FinanceTracker")
        save = settings.value("save_filters", False, type=bool)
        if save:
            raw = settings.value("history_filters", None)
            try:
                data = json.loads(raw) if raw else None
            except Exception:
                data = None
            if data:
                self.type_filter.setCurrentText(data.get("type", "All"))
                self._load_category_filter()
                self.category_filter.setCurrentText(data.get("category", "All"))
                
                current_display_format = settings.value("date_format", "(YYYY-MM-DD) | Year-Month-Day")
                display_format = map_display_format(current_display_format)
                self.start_date.setDisplayFormat(display_format)
                self.end_date.setDisplayFormat(display_format)
                self.start_date.setDate(QDate.fromString(data.get("start", QDate.currentDate().addMonths(-1).toString("yyyy-MM-dd")), "yyyy-MM-dd"))
                self.end_date.setDate(QDate.fromString(data.get("end", QDate.currentDate().toString("yyyy-MM-dd")), "yyyy-MM-dd"))
                self.note_search.setText(data.get("note", ""))
        self._load_category_filter()
        self.update_page_size_from_viewport()
        self.load_page()

    # ---------- UI ----------
    def _build_ui(self):
        root = QVBoxLayout(self)
        # Filters row
        fr = QHBoxLayout()
        self.type_filter = QComboBox()
        self.type_filter.addItems(["All", "Income", "Expense"])
        self.type_filter.currentIndexChanged.connect(self._load_category_filter)
        self.type_filter.currentIndexChanged.connect(self._apply_filters)
        self.category_filter = QComboBox()
        self.category_filter.addItem("All")

        settings = QSettings("Azralithia", "FinanceTracker")
        current_display_format = settings.value("date_format", "(YYYY-MM-DD) | Year-Month-Day")
        display_format = map_display_format(current_display_format)
        
        self.start_date = QDateEdit()
        self.start_date.setCalendarPopup(True)
        self.start_date.setDisplayFormat(display_format) 
        self.start_date.setDate(QDate.currentDate().addMonths(-1))
        
        self.end_date = QDateEdit()
        self.end_date.setCalendarPopup(True)
        self.end_date.setDisplayFormat(display_format)
        self.end_date.setDate(QDate.currentDate())

        self.note_search = QLineEdit()
        self.note_search.setPlaceholderText("Search notes…")

        self.apply_btn = QPushButton("Apply")
        self.reset_btn = QPushButton("Reset")

        for w in (QLabel("Type:"), self.type_filter,
                  QLabel("Category:"), self.category_filter,
                  QLabel("From:"), self.start_date,
                  QLabel("To:"), self.end_date,
                  self.note_search, self.apply_btn, self.reset_btn):
            fr.addWidget(w)
        fr.addStretch()
        root.addLayout(fr)

        # Table
        self.table = QTableWidget()
        self.table.setColumnCount(7)
        self.table.setHorizontalHeaderLabels(
            ["ID", "Date", "Type", "Category", "Amount", "Note", "Actions"]
        )
        self.table.horizontalHeader().setStretchLastSection(False)
        header = self.table.horizontalHeader()
        for i in range(6):
            header.setSectionResizeMode(i, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(6, QHeaderView.ResizeMode.Fixed)
        self.table.setColumnWidth(6, 100)
        self.table.verticalHeader().setDefaultSectionSize(36)
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self.table.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self.table.cellClicked.connect(self._handle_cell_click)
        root.addWidget(self.table)

        # Pager
        pr = QHBoxLayout()
        self.prev_btn = QPushButton("⬅️")
        self.next_btn = QPushButton("➡️")
        for b in (self.prev_btn, self.next_btn):
            b.setFixedSize(BUTTON_WIDTH, BUTTON_HEIGHT)
            b.setCursor(Qt.CursorShape.PointingHandCursor)
        self.prev_btn.setToolTip("Previous 10")
        self.next_btn.setToolTip("Next 10")
        self.page_label = QLabel("Page 1 / 1")

        pr.addStretch()
        pr.addWidget(self.prev_btn)
        pr.addWidget(self.next_btn)
        pr.addWidget(self.page_label)
        root.addLayout(pr)

        # Hooks
        self.apply_btn.clicked.connect(self._apply_filters)
        self.reset_btn.clicked.connect(self._reset_filters)
        self.prev_btn.clicked.connect(self._prev_page)
        self.next_btn.clicked.connect(self._next_page)
        
    def _handle_cell_click(self, row, column):
        transaction_id_item = self.table.item(row, 0)
        if not transaction_id_item:
            return
        
        transaction_id = int(transaction_id_item.text())
        
        conn = sqlite3.connect(self.db_path)
        cur = conn.cursor()
        cur.execute("SELECT date, type, category, amount, note FROM transactions WHERE id=?", (transaction_id,))
        data = cur.fetchone()
        conn.close()
        if not data:
            return 
        date, ttype, category, amount, note = data
        dialog = TransactionEditDialog(transaction_id, self.db_path, self)

        # If the user clicks on data from the table, opens up Editing focus on that data
        if column == 1: 
            dialog.date_edit.setDate(QDate.fromString(date, "yyyy-MM-dd"))
            dialog.date_edit.setFocus()
        elif column == 2: 
            dialog.type_combo.setCurrentText(ttype.title())
            dialog.type_combo.setFocus()
        elif column == 3: 
            dialog.category_edit.setText(category)
            dialog.category_edit.setFocus()
        elif column == 4: 
            dialog.amount_edit.setText(str(amount))
            dialog.amount_edit.setFocus()
        elif column == 5: 
            dialog.note_edit.setText(note)
            dialog.note_edit.setFocus()
        else: 
            pass
        if dialog.exec():
            self.load_page()
            self.data_changed.emit()

    def update_page_size_from_viewport(self):
        row_h = self.table.verticalHeader().defaultSectionSize() or 36
        viewport_h = self.table.viewport().height()
        if viewport_h <= 0:
            return False  
        new_page_size = max(1, viewport_h // row_h)
        if new_page_size != self.page_size:
            self.page_size = int(new_page_size)
            self.current_page = max(0, min(self.current_page, max(0, (self.total_rows - 1) // self.page_size)))
            return True
        return False

    # ---------- Filters & Paging ----------
    def _reset_filters(self):
        self.type_filter.setCurrentIndex(0)
        self.category_filter.setCurrentIndex(0)
        self.start_date.setDate(QDate.currentDate().addMonths(-1))
        self.end_date.setDate(QDate.currentDate())
        self.note_search.clear()
        self.current_page = 0
        self.load_page()

    def _apply_filters(self):
        self.current_page = 0
        self.load_page()

    def _prev_page(self):
        if self.current_page > 0:
            self.current_page -= 1
            self.load_page()

    def _next_page(self):
        max_page = max(0, (self.total_rows - 1) // self.page_size)
        if self.current_page < max_page:
            self.current_page += 1
            self.load_page()

    def _build_where_and_params(self):
        conds, params = [], []

        t = self.type_filter.currentText()
        if t != "All":
            conds.append("lower(type) = ?")
            params.append(t.lower())

        c = self.category_filter.currentText()
        if c != "All":
            conds.append("lower(category) = ?")
            params.append(c.lower())

        sd = self.start_date.date().toString("yyyy-MM-dd")
        ed = self.end_date.date().toString("yyyy-MM-dd")
        if sd:
            conds.append("date >= ?")
            params.append(sd)
        if ed:
            conds.append("date <= ?")
            params.append(ed)

        q = self.note_search.text().strip()
        if q:
            conds.append("note LIKE ?")
            params.append(f"%{q}%")

        where = ("WHERE " + " AND ".join(conds)) if conds else ""
        return where, params

    def _load_category_filter(self):
        settings = QSettings("Azralithia", "FinanceTracker")
        stored = settings.value("categories", None)
        defaults = {
            "Income": ["Salary", "Gift", "Bonus", "Other"],
            "Expense": ["Food", "Rent", "Utilities", "Transport", "Misc"]
        }
        try:
            cats_map = json.loads(stored) if stored else defaults
        except Exception:
            cats_map = defaults

        t = self.type_filter.currentText()
        items = []
        if t == "All":
            seen = set()
            for lst in cats_map.values():
                for c in lst:
                    if c and c.lower() not in seen:
                        seen.add(c.lower())
                        items.append(c.title())
        else:
            items = [c.title() for c in cats_map.get(t, [])]

        self.category_filter.blockSignals(True)
        current = self.category_filter.currentText()
        self.category_filter.clear()
        self.category_filter.addItem("All")
        for c in items:
            self.category_filter.addItem(c)
        idx = self.category_filter.findText(current)
        if idx != -1:
            self.category_filter.setCurrentIndex(idx)
        self.category_filter.blockSignals(False)

    def load_page(self):
        where, params = self._build_where_and_params()
        conn = sqlite3.connect(self.db_path)
        cur = conn.cursor()
        cur.execute(f"SELECT COUNT(*) FROM transactions {where}", params)
        self.total_rows = cur.fetchone()[0]
        limit = self.page_size
        offset = self.current_page * self.page_size
        cur.execute(
            f"""SELECT id, date, type, category, amount, COALESCE(note,'')
                FROM transactions
                {where}
                ORDER BY date DESC, id DESC
                LIMIT ? OFFSET ?""",
            (*params, limit, offset),
        )
        rows = cur.fetchall()
        conn.close()
        self.table.setRowCount(len(rows))
        for r, (rid, date, t, cat, amt, note) in enumerate(rows):
            self.table.setItem(r, 0, QTableWidgetItem(str(rid)))
            self.table.setItem(r, 1, QTableWidgetItem(format_date_for_display(date)))
            self.table.setItem(r, 2, QTableWidgetItem(t.title()))
            self.table.setItem(r, 3, QTableWidgetItem((cat or "").title()))
            self.table.setItem(r, 4, QTableWidgetItem(f"{amt:.2f}"))
            self.table.setItem(r, 5, QTableWidgetItem(note))

            btns_widget = QWidget()
            btns_layout = QHBoxLayout(btns_widget)
            btns_layout.setContentsMargins(0, 0, 0, 0)
            btns_layout.setSpacing(6)
            if self.main_window and rid in self.main_window.pending_delete_transactions:
                entry = self.main_window.pending_delete_transactions[rid]
                labels_map = entry.setdefault("countdown_labels", {})

                if "history" in labels_map and labels_map["history"] is not None:
                    countdown_label = labels_map["history"]
                else:
                    countdown_label = QLabel(str(entry["countdown"]))
                    countdown_label.setFixedWidth(20)
                    countdown_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
                    countdown_label.setStyleSheet("font-weight: bold; color: red;")
                    labels_map["history"] = countdown_label

                buttons_map = entry.setdefault("undo_buttons", {})
                undo_btn = QPushButton("Undo")
                undo_btn.setCursor(Qt.CursorShape.PointingHandCursor)
                undo_btn.setToolTip("Undo Delete")
                undo_btn.clicked.connect(lambda _, _rid=rid: self.main_window.undo_delete(_rid))
                buttons_map["history"] = undo_btn
                inline_undo_widget = QWidget()
                inline_undo_layout = QHBoxLayout(inline_undo_widget)
                inline_undo_layout.setContentsMargins(0, 0, 0, 0)
                inline_undo_layout.setSpacing(2)
                inline_undo_layout.addWidget(undo_btn)
                inline_undo_layout.addWidget(countdown_label)
                inline_undo_layout.addStretch()
                btns_layout.addWidget(inline_undo_widget)
            else:
                edit_btn = QPushButton(" ✏️")
                delete_btn = QPushButton(" 🗑️")
                for b, tip in ((edit_btn, "Edit"), (delete_btn, "Delete")):
                    b.setFixedSize(BUTTON_WIDTH, BUTTON_HEIGHT)
                    b.setCursor(Qt.CursorShape.PointingHandCursor)
                    b.setToolTip(tip)
                edit_btn.clicked.connect(lambda _, _rid=rid: self.edit_transaction(_rid))
                delete_btn.clicked.connect(lambda _, _rid=rid, _date=date, _type=t, _cat=cat, _amt=amt, _note=note:
                    self.main_window.mark_transaction_for_deletion(_rid, (_rid, _date, _type, _cat, _amt, _note))
                )
                btns_layout.addWidget(edit_btn)
                btns_layout.addWidget(delete_btn)

            btns_layout.addStretch()
            self.table.setCellWidget(r, 6, btns_widget)
            self.table.setRowHeight(r, self.table.verticalHeader().defaultSectionSize())
        max_page = max(0, (self.total_rows - 1) // self.page_size)
        self.page_label.setText(f"Page {self.current_page + 1} / {max_page + 1 if self.total_rows else 1}")
        self.prev_btn.setEnabled(self.current_page > 0)
        self.next_btn.setEnabled(self.current_page < max_page)

    def edit_transaction(self, transaction_id: int):
        dialog = TransactionEditDialog(transaction_id, self.db_path, self)
        if dialog.exec():
            self.load_page()
            self.data_changed.emit()

class SettingsPage(QWidget):
    def __init__(self):
        super().__init__()
        self.settings = QSettings("Azralithia", "FinanceTracker")
        self._build_ui()
        self._load_settings()
    
    def _build_ui(self):
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(20, 20, 20, 20)
        main_layout.setSpacing(15)

        # Title
        title_label = QLabel("⚙️ Settings")
        title_label.setStyleSheet("font-size: 24px; font-weight: bold;")
        main_layout.addWidget(title_label, alignment=Qt.AlignmentFlag.AlignLeft)

        # Container for all rows
        self.settings_container = QVBoxLayout()
        self.settings_container.setContentsMargins(0, 0, 0, 0)
        self.settings_container.setSpacing(0)
        main_layout.addLayout(self.settings_container)

        # Rows
        self.date_format_combo = QComboBox()
        self.date_format_combo.addItems(["(DD-MM-YYYY) | Day-Month-Year", "(MM-DD-YYYY) | Month-Day-Year",
                                         "(YYYY-MM-DD) | Year-Month-Day", "(YYYY-DD-MM) | Year-Day-Month"
        ])
        self._create_setting_row("Date Format:", self.date_format_combo)
        self.date_format_combo.currentTextChanged.connect(lambda v: self.settings.setValue("date_format", v))
        self.save_filters_switch = self._create_toggle_setting("Save Filters on Exit:", "save_filters", False)
        self.remember_export_filters_switch = self._create_toggle_setting("Remember Export Filters:", "remember_export_filters", False)
        self.confirm_delete_switch = self._create_toggle_setting("Confirm Delete Transactions:", "confirm_delete", True)
        self.show_undo_on_delete_switch = self._create_toggle_setting("Show Undo Option on Delete:", "show_undo_on_delete", True)
        self.show_undo_confirmation_switch = self._create_toggle_setting("Show Undo Cancelled Confirmation:", "show_undo_confirmation", True)
        self.load_last_page_switch = self._create_toggle_setting("Load Last Page Viewed:", "load_last_page", False)

        dir_edit = QLineEdit()
        dir_edit.setReadOnly(True)
        browse_btn = QPushButton("Browse...")
        browse_btn.clicked.connect(lambda: self._browse_default_directory(dir_edit))
        dir_widget = QWidget()
        dir_layout = QHBoxLayout(dir_widget)
        dir_layout.setContentsMargins(0, 0, 0, 0)
        dir_layout.addWidget(dir_edit)
        dir_layout.addWidget(browse_btn)
        self._create_setting_row("Default Export Dir:", dir_widget)
        self.default_dir_edit = dir_edit
        self.default_dir_edit.textChanged.connect(lambda v: self.settings.setValue("default_export_dir", v))

        self.logging_level_combo = QComboBox()
        self.logging_level_combo.addItems(["DEBUG", "INFO", "WARNING", "ERROR", "CRITICAL"])
        self._create_setting_row("Logging Level:", self.logging_level_combo)
        self.logging_level_combo.currentTextChanged.connect(self._update_logging_level)

        main_layout.addStretch()

    # -=- Helpers -=-
    def _create_setting_row(self, label_text, widget):
        row_frame = QFrame()
        row_frame.setObjectName("settings-row-frame")
        row_layout = QHBoxLayout(row_frame)
        row_layout.setContentsMargins(0, 0, 0, 0)
        row_layout.setSpacing(0)

        label = QLabel(label_text)
        label.setObjectName("settings-label")
        label.setMinimumWidth(200)

        value_container = QWidget()
        value_container.setObjectName("settings-value-widget")
        value_layout = QHBoxLayout(value_container)
        value_layout.setContentsMargins(5, 5, 5, 5)
        value_layout.addWidget(widget)
        value_layout.addStretch()

        row_layout.addWidget(label)
        row_layout.addWidget(value_container)
        self.settings_container.addWidget(row_frame)
        return label, widget

    def _create_toggle_setting(self, label_text, setting_key, default=False):
        switch = ToggleSwitch()
        switch._emoji = ""
        self._create_setting_row(label_text, switch)
        switch.setChecked(self.settings.value(setting_key, default, type=bool))
        switch.toggled.connect(lambda v: self.settings.setValue(setting_key, bool(v)))
        return switch

    def _browse_default_directory(self, line_edit: QLineEdit):
        directory = QFileDialog.getExistingDirectory(
            self, "Select Default Directory", line_edit.text(),
            QFileDialog.Option.ShowDirsOnly | QFileDialog.Option.DontUseNativeDialog
        )
        if directory:
            line_edit.setText(directory)

    def _update_logging_level(self, level_str: str):
        level_map = {
            "DEBUG": logging.DEBUG,
            "INFO": logging.INFO,
            "WARNING": logging.WARNING,
            "ERROR": logging.ERROR,
            "CRITICAL": logging.CRITICAL
        }
        new_level = level_map.get(level_str, logging.INFO)
        logger.setLevel(new_level)
        self.settings.setValue("logging_level", level_str)
        logger.info(f"Logging level set to {level_str}")

    def _load_settings(self):
        self.date_format_combo.setCurrentText(
            self.settings.value("date_format", "(YYYY-MM-DD) | Year-Month-Day")
        )
        self.default_dir_edit.setText(self.settings.value("default_export_dir", ""))
        self.logging_level_combo.setCurrentText(self.settings.value("logging_level", "INFO"))

   
# --- Functionality ---
class TransactionEditDialog(QDialog):
    def __init__(self, transaction_id, db_path, parent=None):
        super().__init__(parent)
        self.db_path = db_path
        self.transaction_id = transaction_id
        self.setWindowTitle("Edit Transaction")
        self.resize(400, 220)
        self.settings = QSettings("Azralithia", "FinanceTracker")

        self._build_ui()
        self.load_data()

    def _build_ui(self):
        layout = QFormLayout(self)

        # Date
        self.date_edit = QDateEdit()
        self.date_edit.setCalendarPopup(True)
        display_format = map_display_format(
            self.settings.value("date_format", "(YYYY-MM-DD) | Year-Month-Day")
        )
        self.date_edit.setDisplayFormat(display_format)

        # Type and Category
        self.type_combo = QComboBox()
        self.type_combo.addItems(["Income", "Expense"])
        self.type_combo.currentIndexChanged.connect(self._load_categories_for_type)
        self.category_combo = QComboBox()

        # Amount & Note
        self.amount_edit = QLineEdit()
        self.note_edit = QLineEdit()

        # Rows
        layout.addRow("Date:", self.date_edit)
        layout.addRow("Type:", self.type_combo)
        layout.addRow("Category:", self.category_combo)
        layout.addRow("Amount:", self.amount_edit)
        layout.addRow("Note:", self.note_edit)

        # Buttons
        btn_box = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Save | QDialogButtonBox.StandardButton.Cancel
        )
        btn_box.accepted.connect(self.save_changes)
        btn_box.rejected.connect(self.reject)
        layout.addRow(btn_box)

    def _load_categories_for_type(self):
        selected_type = self.type_combo.currentText()
        stored = self.settings.value("categories", None)
        defaults = {
            "Income": ["Salary", "Gift", "Bonus", "Other"],
            "Expense": ["Food", "Rent", "Utilities", "Transport", "Misc"]
        }
        try:
            cats_map = json.loads(stored) if stored else defaults
        except Exception:
            cats_map = defaults

        categories = cats_map.get(selected_type, [])
        self.category_combo.blockSignals(True)
        self.category_combo.clear()
        self.category_combo.addItems([cat.title() for cat in categories])
        self.category_combo.blockSignals(False)

    def load_data(self):
        conn = sqlite3.connect(self.db_path)
        cur = conn.cursor()
        cur.execute(
            "SELECT date, type, category, amount, note FROM transactions WHERE id=?",
            (self.transaction_id,)
        )
        row = cur.fetchone()
        conn.close()
        if not row:
            return

        date, ttype, category, amount, note = row
        self.date_edit.setDate(QDate.fromString(date, "yyyy-MM-dd"))
        self.type_combo.setCurrentText(ttype.title())
        self._load_categories_for_type()

        categories = [self.category_combo.itemText(i) for i in range(self.category_combo.count())]
        if category and category.title() in categories:
            self.category_combo.setCurrentText(category.title())
        else:
            fallback = "Other" if "Other" in categories else categories[0] if categories else "Other"
            if "Other" not in categories:
                self.category_combo.addItem("Other")
            self.category_combo.setCurrentText(fallback)

        self.amount_edit.setText(str(amount))
        self.note_edit.setText(note or "")

    def save_changes(self):
        date = self.date_edit.date().toString("yyyy-MM-dd")
        ttype = self.type_combo.currentText().lower()
        category = self.category_combo.currentText().lower()

        try:
            amount = float(self.amount_edit.text())
        except ValueError:
            QMessageBox.warning(self, "Invalid Input", "Amount must be a number.")
            return

        note = self.note_edit.text()

        conn = sqlite3.connect(self.db_path)
        cur = conn.cursor()
        cur.execute(
            "UPDATE transactions SET date=?, type=?, category=?, amount=?, note=? WHERE id=?",
            (date, ttype, category, amount, note, self.transaction_id),
        )
        conn.commit()
        conn.close()
        self.accept()

class ExportOptionsDialog(QDialog):
    def __init__(self, db_path, parent=None):
        super().__init__(parent)
        self.db_path = db_path
        self.settings = QSettings("Azralithia", "FinanceTracker")
        self.current_filters = {}
        self.filter_scope = "full"

        self.setWindowTitle("Export Transactions")
        self.resize(500, 300)

        self._build_ui()
        self._load_initial_filters()
        self._update_filter_summary()

    def _build_ui(self):
        main_layout = QVBoxLayout(self)
        main_layout.setSpacing(15)

        scope_group = QGroupBox("Export Scope")
        scope_layout = QVBoxLayout(scope_group)
        self.full_db_radio = QRadioButton("Export Full Database")
        self.filtered_radio = QRadioButton("Export Filtered Transactions")
        scope_layout.addWidget(self.full_db_radio)
        scope_layout.addWidget(self.filtered_radio)
        main_layout.addWidget(scope_group)

        self.filter_group = QGroupBox("Filters for Export")
        filter_layout = QVBoxLayout(self.filter_group)
        self.filter_summary_label = QLabel("No filters applied.")
        self.filter_summary_label.setWordWrap(True)
        self.edit_filters_btn = QPushButton("⚙️ Edit Filters")
        self.edit_filters_btn.setFixedSize(120, 30)
        filter_layout.addWidget(self.filter_summary_label)
        filter_layout.addWidget(self.edit_filters_btn, alignment=Qt.AlignmentFlag.AlignRight)
        main_layout.addWidget(self.filter_group)

        format_group = QGroupBox("Export Format")
        format_layout = QHBoxLayout(format_group)
        self.format_combo = QComboBox()
        self.format_combo.addItems(["CSV", "Excel (XLSX)", "JSON"])
        format_layout.addWidget(self.format_combo)
        format_layout.addStretch()
        main_layout.addWidget(format_group)

        button_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        button_box.button(QDialogButtonBox.StandardButton.Ok).setText("Export")
        main_layout.addWidget(button_box)

        self.full_db_radio.toggled.connect(self._toggle_filter_scope)
        self.filtered_radio.toggled.connect(self._toggle_filter_scope)
        self.edit_filters_btn.clicked.connect(self._open_filter_editor)
        button_box.accepted.connect(self._perform_export)
        button_box.rejected.connect(self.reject)

        self.full_db_radio.setChecked(True)
        self._toggle_filter_scope()

    def _load_initial_filters(self):
        remember = self.settings.value("remember_export_filters", False, type=bool)
        raw = self.settings.value("history_filters", None)
        if remember and raw:
            try:
                self.current_filters = json.loads(raw)
                self.filtered_radio.setChecked(True)
            except json.JSONDecodeError:
                self.current_filters = {}
        self.current_filters.setdefault("type", "All")
        self.current_filters.setdefault("category", "All")
        self.current_filters.setdefault("start", QDate.currentDate().addMonths(-1).toString("yyyy-MM-dd"))
        self.current_filters.setdefault("end", QDate.currentDate().toString("yyyy-MM-dd"))
        self.current_filters.setdefault("note", "")

    def _toggle_filter_scope(self):
        self.filter_scope = "full" if self.full_db_radio.isChecked() else "filtered"
        self.filter_group.setEnabled(self.filter_scope == "filtered")
        self._update_filter_summary()

    def _update_filter_summary(self):
        if self.filter_scope == "full":
            self.filter_summary_label.setText("Exporting all transactions.")
            return

        summary = []
        if self.current_filters.get("type") != "All":
            summary.append(f"Type: {self.current_filters['type']}")
        if self.current_filters.get("category") != "All":
            summary.append(f"Category: {self.current_filters['category']}")
        start, end = self.current_filters.get("start"), self.current_filters.get("end")
        summary.append(f"Date Range: {format_date_for_display(start)} to {format_date_for_display(end)}")
        note = self.current_filters.get("note")
        if note:
            summary.append(f"Note contains: '{note}'")

        self.filter_summary_label.setText(
            "Filters: " + "; ".join(summary) if summary else
            "No specific filters applied (all types/categories, default date range)."
        )

    def _open_filter_editor(self):
        dialog = ExportFilterDialog(self.current_filters, self.db_path, self)
        if dialog.exec():
            self.current_filters = dialog.get_filters()
            self._update_filter_summary()

    def _perform_export(self):
        format_ext = {"CSV": "csv", "Excel (XLSX)": "xlsx", "JSON": "json"}
        ext = format_ext.get(self.format_combo.currentText(), "txt")

        file_name, _ = QFileDialog.getSaveFileName(self, "Save Export File", self.settings.value("default_export_dir", ""),
            f"*.{ext}", options=QFileDialog.Option.DontUseNativeDialog
        )

        if not file_name:
            return
        if not file_name.lower().endswith(f".{ext}"):
            file_name += f".{ext}"

        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()

            if self.filter_scope == "full":
                cursor.execute("SELECT date, type, category, amount, note FROM transactions ORDER BY date DESC")
            else:
                conds, params = [], []
                t, c, sd, ed, q = (self.current_filters.get(k) for k in ["type", "category", "start", "end", "note"])
                if t and t != "All": conds.append("lower(type)=?"); params.append(t.lower())
                if c and c != "All": conds.append("lower(category)=?"); params.append(c.lower())
                if sd: conds.append("date>=?"); params.append(sd)
                if ed: conds.append("date<=?"); params.append(ed)
                if q: conds.append("note LIKE ?"); params.append(f"%{q}%")
                where = ("WHERE " + " AND ".join(conds)) if conds else ""
                cursor.execute(f"SELECT date, type, category, amount, note FROM transactions {where} ORDER BY date DESC", params)

            rows = cursor.fetchall()
            conn.close()
            headers = ["Date", "Type", "Category", "Amount", "Note"]
            fmt = self.format_combo.currentText()
            if fmt == "CSV": self._export_to_csv(file_name, headers, rows)
            elif fmt == "Excel (XLSX)": self._export_to_excel(file_name, headers, rows)
            elif fmt == "JSON": self._export_to_json(file_name, headers, rows)
            else: QMessageBox.warning(self, "Export Error", f"Unsupported format: {fmt}"); return

            QMessageBox.information(self, "Export Successful", f"Data exported to {file_name}")
            self.accept()

        except Exception as e:
            QMessageBox.critical(self, "Export Error", f"An error occurred: {e}")
            logger.error("Export failed", exc_info=True)

    # Export helpers
    def _export_to_csv(self, file_name, headers, rows):
        import csv
        with open(file_name, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            writer.writerows(rows)

    def _export_to_excel(self, file_name, headers, rows):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Transactions"
        ws.append(headers)
        for r in rows: ws.append(r)
        wb.save(file_name)

    def _export_to_json(self, file_name, headers, rows):
        data = [{headers[i].lower(): val for i, val in enumerate(row)} for row in rows]
        with open(file_name, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=4, ensure_ascii=False)

class ExportFilterDialog(QDialog):
    def __init__(self, initial_filters: dict, db_path: str, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Set Export Filters")
        self.db_path = db_path
        self.settings = QSettings("Azralithia", "FinanceTracker")
        self._filters = initial_filters.copy()

        self._build_ui()
        self._load_category_filter()
        self._set_initial_values()

    def _build_ui(self):
        layout = QFormLayout(self)

        self.type_filter = QComboBox()
        self.type_filter.addItems(["All", "Income", "Expense"])
        self.type_filter.currentIndexChanged.connect(self._load_category_filter)
        self.category_filter = QComboBox()
        self.category_filter.addItem("All")

        display_format = map_display_format(self.settings.value("date_format", "(YYYY-MM-DD) | Year-Month-Day"))
        self.start_date = QDateEdit(calendarPopup=True)
        self.start_date.setDisplayFormat(display_format)
        self.end_date = QDateEdit(calendarPopup=True)
        self.end_date.setDisplayFormat(display_format)

        self.note_search = QLineEdit()
        self.note_search.setPlaceholderText("Search notes…")

        layout.addRow("Type:", self.type_filter)
        layout.addRow("Category:", self.category_filter)
        layout.addRow("From Date:", self.start_date)
        layout.addRow("To Date:", self.end_date)
        layout.addRow("Note Search:", self.note_search)

        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Apply | QDialogButtonBox.StandardButton.Cancel)
        buttons.accepted.connect(self._apply_and_accept)
        buttons.rejected.connect(self.reject)
        layout.addRow(buttons)

    def _set_initial_values(self):
        self.type_filter.setCurrentText(self._filters.get("type", "All"))
        self.category_filter.setCurrentText(self._filters.get("category", "All"))
        self.start_date.setDate(QDate.fromString(self._filters.get("start", QDate.currentDate().addMonths(-1).toString("yyyy-MM-dd")), "yyyy-MM-dd"))
        self.end_date.setDate(QDate.fromString(self._filters.get("end", QDate.currentDate().toString("yyyy-MM-dd")), "yyyy-MM-dd"))
        self.note_search.setText(self._filters.get("note", ""))

    def _load_category_filter(self):
        stored = self.settings.value("categories", None)
        defaults = {"Income": ["Salary", "Gift", "Bonus", "Other"], "Expense": ["Food", "Rent", "Utilities", "Transport", "Misc"]}
        try: cats_map = json.loads(stored) if stored else defaults
        except Exception: cats_map = defaults

        t = self.type_filter.currentText()
        items = []
        if t == "All":
            seen = set()
            for lst in cats_map.values():
                for c in lst:
                    if c and c.lower() not in seen:
                        seen.add(c.lower())
                        items.append(c.title())
        else:
            items = [c.title() for c in cats_map.get(t, [])]

        self.category_filter.blockSignals(True)
        current = self.category_filter.currentText()
        self.category_filter.clear()
        self.category_filter.addItem("All")
        for c in items: self.category_filter.addItem(c)
        if current in [self.category_filter.itemText(i) for i in range(self.category_filter.count())]:
            self.category_filter.setCurrentText(current)
        else:
            self.category_filter.setCurrentText("All")
        self.category_filter.blockSignals(False)

    def _apply_and_accept(self):
        self._filters.update({
            "type": self.type_filter.currentText(),
            "category": self.category_filter.currentText(),
            "start": self.start_date.date().toString("yyyy-MM-dd"),
            "end": self.end_date.date().toString("yyyy-MM-dd"),
            "note": self.note_search.text().strip()
        })
        self.accept()

    def get_filters(self) -> dict:
        return self._filters

class ImportOptionsDialog(QDialog):
    def __init__(self, db_path, parent=None):
        super().__init__(parent)
        self.db_path = db_path
        self.setWindowTitle("Import Transactions")
        self.resize(500, 250)
        self._build_ui()

    def _build_ui(self):
        layout = QVBoxLayout(self)

        # File selection
        file_layout = QHBoxLayout()
        self.file_line_edit = QLineEdit()
        self.file_line_edit.setReadOnly(True)
        self.browse_btn = QPushButton("Browse…")
        file_layout.addWidget(QLabel("Import File:"))
        file_layout.addWidget(self.file_line_edit)
        file_layout.addWidget(self.browse_btn)
        layout.addLayout(file_layout)

        # Import mode
        mode_group = QGroupBox("Import Mode")
        mode_layout = QVBoxLayout(mode_group)
        self.override_radio = QRadioButton("Override Existing Database (Replace all data)")
        self.add_radio = QRadioButton("Add to Database (Merge)")
        mode_layout.addWidget(self.override_radio)
        mode_layout.addWidget(self.add_radio)
        layout.addWidget(mode_group)

        # Add mode options
        self.add_options_group = QGroupBox("Add Mode Options")
        add_options_layout = QVBoxLayout(self.add_options_group)
        self.add_all_radio = QRadioButton("Add all transactions (allow duplicates)")
        self.add_missing_radio = QRadioButton("Add only missing transactions (skip duplicates)")
        add_options_layout.addWidget(self.add_all_radio)
        add_options_layout.addWidget(self.add_missing_radio)
        layout.addWidget(self.add_options_group)

        # Buttons
        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        layout.addWidget(buttons)

        # Connections
        self.browse_btn.clicked.connect(self._browse_file)
        self.override_radio.toggled.connect(self._update_add_options_enabled)
        buttons.accepted.connect(self._import)
        buttons.rejected.connect(self.reject)

        # Defaults
        self.override_radio.setChecked(True)
        self.add_all_radio.setChecked(True)
        self._update_add_options_enabled()

    def _browse_file(self):
        file_name, _ = QFileDialog.getOpenFileName(self, "Select Import File", "",
                                                   "CSV Files (*.csv);;Excel Files (*.xlsx);;JSON Files (*.json)",
                                                   options=QFileDialog.Option.DontUseNativeDialog) 
        if file_name:
            self.file_line_edit.setText(file_name)

    def _update_add_options_enabled(self):
        enabled = self.add_radio.isChecked() 
        self.add_options_group.setEnabled(enabled)

    def _import(self):
        file_path = self.file_line_edit.text()
        if not file_path:
            QMessageBox.warning(self, "No File Selected", "Please select a file to import.")
            return

        try:
            transactions = self._load_transactions_from_file(file_path)
        except Exception as e:
            QMessageBox.critical(self, "Import Error", f"Failed to load transactions: {e}")
            return

        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        try:
            if self.override_radio.isChecked():
                cursor.execute("DELETE FROM transactions")
                self._insert_transactions(cursor, transactions)
                conn.commit()
                QMessageBox.information(self, "Import Successful", f"Database overridden with {len(transactions)} transactions.")
            else:
                if self.add_all_radio.isChecked():
                    self._insert_transactions(cursor, transactions)
                    conn.commit()
                    QMessageBox.information(self, "Import Successful", f"Added {len(transactions)} transactions (duplicates allowed).")
                else:
                    added_count = 0
                    for t in transactions:
                        if not self._transaction_exists(cursor, t):
                            self._insert_transaction(cursor, t)
                            added_count += 1
                    conn.commit()
                    QMessageBox.information(self, "Import Successful", f"Added {added_count} new transactions (duplicates skipped).")
        except Exception as e:
            QMessageBox.critical(self, "Import Error", f"Failed during import: {e}")
            conn.rollback()
        finally:
            conn.close()

        self.accept()

    def _load_transactions_from_file(self, file_path):
        ext = os.path.splitext(file_path)[1].lower().replace(".", "")
        transactions = []
        if ext == "csv":
            with open(file_path, newline='', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    transactions.append(self._row_to_transaction(row))
        elif ext == "xlsx":
            wb = openpyxl.load_workbook(file_path)
            sheet = wb.active
            headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_dict = dict(zip(headers, row))
                transactions.append(self._row_to_transaction(row_dict))
        elif ext == "json":
            with open(file_path, encoding='utf-8') as f:
                data = json.load(f)
                for item in data:
                    transactions.append(self._row_to_transaction(item))
        else:
            raise RuntimeError("Unsupported file format.")
        return transactions

    def _row_to_transaction(self, row):
        return {
            "date": row.get("Date") or row.get("date"),
            "type": row.get("Type") or row.get("type"),
            "category": row.get("Category") or row.get("category"),
            "amount": float(row.get("Amount") or row.get("amount") or 0),
            "note": row.get("Note") or row.get("note") or ""
        }

    def _transaction_exists(self, cursor, t):
        cursor.execute("""
            SELECT 1 FROM transactions WHERE
            date = ? AND
            lower(type) = ? AND
            lower(category) = ? AND
            amount = ? AND
            note = ?
            LIMIT 1
        """, (t["date"], t["type"].lower(), t["category"].lower(), t["amount"], t["note"]))
        return cursor.fetchone() is not None

    def _insert_transactions(self, cursor, transactions):
        for t in transactions:
            cursor.execute("""
                INSERT INTO transactions (date, type, category, amount, note)
                VALUES (?, ?, ?, ?, ?)
            """, (t["date"], t["type"], t["category"], t["amount"], t["note"]))

class CategoryEditor(QDialog):
    def __init__(self, categories, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Edit Categories")
        layout = QVBoxLayout(self)

        self.category_list = QListWidget()
        self.category_list.addItems(categories)
        layout.addWidget(self.category_list)

        btn_row = QHBoxLayout()
        add_btn = QPushButton("➕ Add")
        remove_btn = QPushButton("➖ Remove")
        btn_row.addWidget(add_btn)
        btn_row.addWidget(remove_btn)
        layout.addLayout(btn_row)

        save_btn = QPushButton("💾 Save")
        layout.addWidget(save_btn)

        add_btn.clicked.connect(self.add_category)
        remove_btn.clicked.connect(self.remove_category)
        save_btn.clicked.connect(self.accept)

    def add_category(self):
        text, ok = QInputDialog.getText(self, "New Category", "Enter name:")
        if ok:
            text = text.strip()
            if text:
                existing = {self.category_list.item(i).text() for i in range(self.category_list.count())}
                if text not in existing:
                    self.category_list.addItem(text)

    def remove_category(self):
        for item in self.category_list.selectedItems():
            self.category_list.takeItem(self.category_list.row(item))

    def get_categories(self):
        return [self.category_list.item(i).text() for i in range(self.category_list.count())]


# --- Main Window / UI ---
class ToggleSwitch(QCheckBox):
    thumbChanged = pyqtSignal()

    def __init__(self):
        super().__init__()
        self.setCheckable(True)
        self.setChecked(False)          
        self.setCursor(Qt.CursorShape.PointingHandCursor)
        self.setFixedSize(TOGGLE_WIDTH, TOGGLE_HEIGHT)

        self._thumb_x = 2
        self._track_dark = True
        self._emoji = "🌙"

        # Animation
        self._anim = QPropertyAnimation(self, b"thumb_pos", self)
        self._anim.setDuration(TOGGLE_ANIM_DURATION)
        self._anim.setEasingCurve(QEasingCurve.Type.InOutQuad)

        # Hook state → animation & emoji
        self.toggled.connect(self._animate_to_state)
        self.toggled.connect(self._update_emoji)

    def get_thumb_pos(self) -> float:
        return float(self._thumb_x)

    def set_thumb_pos(self, x: float):
        self._thumb_x = float(x)
        self.thumbChanged.emit()
        self.update()

    thumb_pos = pyqtProperty(float, fget=get_thumb_pos, fset=set_thumb_pos, notify=thumbChanged)

    #  -=- Helpers -=-
    def _left_pos(self) -> int:
        return TOGGLE_MARGIN

    def _right_pos(self) -> int:
        thumb_d = self.height() - TOGGLE_PADDING
        return self.width() - thumb_d - TOGGLE_MARGIN

    def _animate_to_state(self, checked: bool):
        start = self._thumb_x
        end = self._right_pos() if checked else self._left_pos()
        self._anim.stop()
        self._anim.setStartValue(start)
        self._anim.setEndValue(end)
        self._anim.start()

    def _update_emoji(self, checked: bool):
        if self._emoji == "":
            return
        self._emoji = "☀️" if checked else "🌙"
        self.update()

    def resizeEvent(self, e):
        self._thumb_x = self._right_pos() if self.isChecked() else self._left_pos()
        super().resizeEvent(e)

    def mousePressEvent(self, event):
        """Make the whole widget clickable; avoid double-toggle by consuming the event."""
        if event.button() == Qt.MouseButton.LeftButton:
            self.toggle()
            event.accept()
            return
        super().mousePressEvent(event)

    def paintEvent(self, event):
        from PyQt6.QtGui import QPainter, QColor, QBrush, QFont
        p = QPainter(self)
        p.setRenderHint(QPainter.RenderHint.Antialiasing)

        rect = self.rect()
        radius = rect.height() / 2

        # Track color depends on theme + state
        if self.isChecked():
            track_color = QColor("#00c853") if self._track_dark else QColor("#1976d2")
        else:
            track_color = QColor("#555") if self._track_dark else QColor("#bbb")

        # Track
        p.setBrush(QBrush(track_color))
        p.setPen(Qt.PenStyle.NoPen)
        p.drawRoundedRect(rect, radius, radius)

        # Thumb
        thumb_d = rect.height() - TOGGLE_PADDING
        thumb_y = (rect.height() - thumb_d) // 2
        thumb_rect = QRect(int(self._thumb_x), int(thumb_y), int(thumb_d), int(thumb_d))

        p.setBrush(QBrush(QColor("#fff")))
        p.drawEllipse(thumb_rect)

        # Emoji inside the thumb (moves with it)
        p.setFont(QFont("Segoe UI Emoji", 11))
        p.setPen(QColor("#000"))
        p.drawText(thumb_rect, Qt.AlignmentFlag.AlignCenter, self._emoji)

        p.end()

class SidebarButton(QPushButton):
    clicked_action = pyqtSignal(str)

    def __init__(self, icon, text, is_subitem=False, start_collapsed=True):
        super().__init__()
        self.icon_only = icon
        self.full_text = f"{icon} {text}"
        self.setCursor(Qt.CursorShape.PointingHandCursor)
        self.setFixedHeight(BUTTON_HEIGHT)
        self.setProperty("subitem", is_subitem)

        self._collapsed = start_collapsed
        self.setText(self.icon_only if start_collapsed else self.full_text)

        self.clicked.connect(lambda: self.clicked_action.emit(text))

    def set_collapsed(self, collapsed: bool):
        if collapsed == self._collapsed:
            return
        self._collapsed = collapsed
        self.setText(self.icon_only if collapsed else self.full_text)

class Submenu(QWidget):
    def __init__(self, buttons):
        super().__init__()
        self.layout = QVBoxLayout(self)
        self.layout.setContentsMargins(0, 0, 0, 0)
        self.layout.setSpacing(0)
        for btn in buttons:
            self.layout.addWidget(btn)

        self.setMaximumHeight(0)
        self.anim = QPropertyAnimation(self, b"maximumHeight")
        self.anim.setDuration(ANIM_DURATION)
        self.anim.setEasingCurve(QEasingCurve.Type.InOutQuad)

    def toggle(self):
        target_height = BUTTON_HEIGHT * self.layout.count() if self.maximumHeight() == 0 else 0
        self.anim.stop()
        self.anim.setStartValue(self.maximumHeight())
        self.anim.setEndValue(target_height)
        self.anim.start()

class Sidebar(QWidget):
    theme_toggled = pyqtSignal(bool)
    action_triggered = pyqtSignal(str)

    def __init__(self):
        super().__init__()
        self.setObjectName("Sidebar")
        self.setAttribute(Qt.WidgetAttribute.WA_StyledBackground, True)
        self.setMaximumWidth(SIDEBAR_COLLAPSED_WIDTH)
        self._collapsed = True  

        # Smooth animation for sidebar width
        self.animation = QPropertyAnimation(self, b"maximumWidth")
        self.animation.setDuration(ANIM_DURATION)
        self.animation.setEasingCurve(QEasingCurve.Type.InOutQuad)
        self.animation.finished.connect(self._on_animation_finished)

        self.layout = QVBoxLayout(self)
        self.layout.setContentsMargins(0, 0, 0, 0)
        self.layout.setSpacing(0)

        self.buttons = []

        # Main buttons
        self.add_button("🏠", "Main")
        self.finance_btn = self.add_button("💰", "Finance Tracker")

        # Submenu
        self.sub_buttons = [
            SidebarButton("📝", "Manage Transactions", is_subitem=True),
            SidebarButton("📊", "Show Summary", is_subitem=True),
            SidebarButton("📜", "Transaction Log", is_subitem=True),
            SidebarButton("📤", "Export", is_subitem=True),
            SidebarButton("📥", "Import", is_subitem=True),
        ]
        self.submenu = Submenu(self.sub_buttons)
        self.layout.addWidget(self.submenu)
        self.add_button("⚙️", "Settings")
        self.layout.addStretch()  

        # Exit + Theme switch row
        bottom_container = QWidget()
        bottom_container.setObjectName("SidebarBottom") 
        bottom_container.setAttribute(Qt.WidgetAttribute.WA_StyledBackground, True)
        bottom_layout = QHBoxLayout(bottom_container)
        bottom_layout.setContentsMargins(5, 5, 5, 5)
        bottom_layout.setSpacing(10)

        self.exit_btn = SidebarButton("🚪", "Exit")
        self.theme_switch = ToggleSwitch()
        self.theme_switch.toggled.connect(self.toggle_theme)
        self.theme_switch.setVisible(False)
        
        bottom_layout.addWidget(self.exit_btn)
        bottom_layout.addWidget(self.theme_switch)

        self.layout.addWidget(bottom_container) 
        self.finance_btn.clicked.connect(self.submenu.toggle)

        # Relay signals
        self.exit_btn.clicked_action.connect(self.action_triggered)
        for btn in self.buttons + self.sub_buttons:
            btn.clicked_action.connect(self.action_triggered)

    def add_button(self, icon, text, sub=False):
        btn = SidebarButton(icon, text, is_subitem=sub, start_collapsed=True)
        self.layout.addWidget(btn)
        self.buttons.append(btn)
        return btn

    def animate_sidebar(self, collapsed: bool):
        if collapsed == self._collapsed:
            return
        self._collapsed = collapsed
        self.animation.stop()
        self.animation.setStartValue(self.width())
        end = SIDEBAR_COLLAPSED_WIDTH if collapsed else SIDEBAR_EXPANDED_WIDTH
        self.animation.setEndValue(end)
        if not collapsed:
            for btn in (*self.buttons, *self.sub_buttons, self.exit_btn):
                btn.set_collapsed(False)
            self.theme_switch.setVisible(True)
        self.animation.start()

    def _on_animation_finished(self):
        if self._collapsed:
            for btn in (*self.buttons, *self.sub_buttons, self.exit_btn):
                btn.set_collapsed(True)
            self.theme_switch.setVisible(False)

    def enterEvent(self, event):
        self.animate_sidebar(collapsed=False)
        super().enterEvent(event)

    def leaveEvent(self, event):
        self.animate_sidebar(collapsed=True)
        super().leaveEvent(event)

    def toggle_theme(self, state):
        self.theme_toggled.emit(bool(state))

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.pending_delete_transactions = {}
        self.delete_countdown_timer = QTimer(self) 
        self.delete_countdown_timer.timeout.connect(self._update_delete_countdowns) 
        self.delete_countdown_timer.start(1000)
        self.setWindowIcon(QIcon(resource_path("assets/icon.png")))
        create_db()

        self.setWindowTitle("Azralithia Finance Tracker")
        self.setMinimumSize(1200, 700)
        self.setGeometry(200, 200, 900, 600)
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        layout = QHBoxLayout(central_widget)
        self.sidebar = Sidebar()
        layout.addWidget(self.sidebar)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)

        self.stack = QStackedWidget()
        self.transactions_page = TransactionsPage(main_window=self)
        self.stack.addWidget(self.transactions_page)
        self.show_summary_tab = SummaryPage(db_path="transactions.db")
        self.stack.addWidget(self.show_summary_tab)
        self.history_page = HistoryPage(db_path="transactions.db", main_window=self) 
        self.stack.addWidget(self.history_page)
        self.main_page = MainPage(self)
        self.main_page.set_page_switcher(self.stack, self.transactions_page)
        self.settings_page = SettingsPage()
        self.stack.addWidget(self.main_page)
        self.stack.addWidget(self.settings_page)
        layout.addWidget(self.stack)
        self.stack.setCurrentWidget(self.main_page)
    
        # Connect sidebar
        self.sidebar.action_triggered.connect(self.handle_action)
        self.sidebar.theme_toggled.connect(self.toggle_theme)
        self.settings = QSettings("Azralithia", "FinanceTracker")
        light = self.settings.value("light_mode", False, type=bool)

        # Sync toggle state with saved setting
        self.sidebar.theme_switch.setChecked(light)
        self.toggle_theme(light)
        self.transactions_page.data_changed.connect(self.refresh_ui)
        self.history_page.data_changed.connect(self.refresh_ui)
        self.transactions_page.data_changed.connect(self.history_page._load_category_filter)
        self.transactions_page.data_changed.connect(self.history_page.load_page)
        self.transactions_page.data_changed.connect(self.main_page.load_summary)
        self.history_page.data_changed.connect(self.main_page.load_summary)

        # Connect setting toggle to QSettings
        self.settings_page.save_filters_switch.toggled.connect(lambda v: self.settings.setValue("save_filters", bool(v)))
        self.settings_page.remember_export_filters_switch.toggled.connect(lambda v: self.settings.setValue("remember_export_filters", bool(v)))
        self.settings_page.date_format_combo.currentTextChanged.connect(lambda v: (self.settings.setValue("date_format", v), self.on_date_format_changed()))
        self.settings_page.confirm_delete_switch.toggled.connect(lambda v: self.settings.setValue("confirm_delete", bool(v)))
        self.settings_page.show_undo_on_delete_switch.toggled.connect(lambda v: self.settings.setValue("show_undo_on_delete", bool(v)))
        self.settings_page.show_undo_confirmation_switch.toggled.connect(lambda v: self.settings.setValue("show_undo_confirmation", bool(v)))
        self.settings_page.load_last_page_switch.toggled.connect(lambda v: self.settings.setValue("load_last_page", bool(v)))
        self.settings_page.default_dir_edit.textChanged.connect(lambda v: self.settings.setValue("default_export_dir", v))
        self.settings_page.logging_level_combo.currentTextChanged.connect(self._update_logging_level_from_settings)

        self._load_last_page_viewed()
        self._update_logging_level_from_settings(self.settings.value("logging_level", "INFO")) 

   # -=- Undo Button -=- 
    def undo_delete(self, rid: int):
        entry = self.pending_delete_transactions.get(rid)
        if not entry:
            QMessageBox.warning(self, "Undo Failed", "No pending deletion found for this transaction.")
            return

        timer = entry.get("timer")
        if timer:
            try:
                timer.stop()
            except Exception:
                pass

        labels_map = entry.get("countdown_labels", {})
        for lbl in list(labels_map.values()):
            try:
                if lbl is not None:
                    lbl.deleteLater()
            except Exception:
                pass
        buttons_map = entry.get("undo_buttons", {})
        for btn in list(buttons_map.values()):
            try:
                if btn is not None:
                    btn.deleteLater()
            except Exception:
                pass

        try:
            del self.pending_delete_transactions[rid]
        except KeyError:
            pass

        logger.info(f"Transaction ID {rid} deletion cancelled.")
        self.refresh_ui()
        show_confirmation = self.settings.value("show_undo_confirmation", True, type=bool)
        if show_confirmation:
            QMessageBox.information(self, "Undo Successful", "Transaction deletion cancelled.")

    def mark_transaction_for_deletion(self, rid: int, transaction_data: tuple):
        # Confirm delete setting
        if self.settings.value("confirm_delete", True, type=bool):
            reply = QMessageBox.question(
                self,
                "Confirm Delete",
                "Are you sure you want to delete this transaction?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
            if reply != QMessageBox.StandardButton.Yes:
                return  

        #  Undo option setting
        show_undo_option = self.settings.value("show_undo_on_delete", True, type=bool)
        if not show_undo_option:
            try:
                conn = sqlite3.connect("transactions.db")
                cursor = conn.cursor()
                cursor.execute("DELETE FROM transactions WHERE id = ?", (rid,))
                conn.commit()
                conn.close()
                logger.info(f"Transaction ID {rid} deleted immediately (undo disabled).")
            except Exception as e:
                logger.error(f"Failed to delete transaction ID {rid}: {e}", exc_info=True)
                QMessageBox.critical(self, "Delete Error", f"An error occurred: {e}")
            finally:
                self.refresh_ui()
            return

        if rid in self.pending_delete_transactions:
            return  

        self.pending_delete_transactions[rid] = {
            "transaction": transaction_data,
            "countdown": 5,
            "undo_button": {},
            "countdown_label": {}
        }
        self.refresh_ui()

    def _update_delete_countdowns(self):
        for rid in list(self.pending_delete_transactions.keys()):
            entry = self.pending_delete_transactions.get(rid)
            if not entry:
                continue

            entry["countdown"] = entry.get("countdown", 5) - 1

            labels_map = entry.get("countdown_labels", {})
            for lbl in list(labels_map.values()):
                try:
                    if lbl is not None:
                        lbl.setText(str(entry["countdown"]))
                except Exception:
                    pass

            if entry["countdown"] <= 0:
                self.finalize_delete(rid)

    def finalize_delete(self, rid: int):
        entry = self.pending_delete_transactions.get(rid)
        if entry:
            timer = entry.get("timer")
            if timer:
                try:
                    timer.stop()
                except Exception:
                    pass
        try:
            conn = sqlite3.connect("transactions.db")
            cursor = conn.cursor()
            cursor.execute("DELETE FROM transactions WHERE id = ?", (rid,))
            conn.commit()
            conn.close()
            logger.info(f"Transaction ID {rid} successfully deleted from DB.")
        except Exception as e:
            logger.error(f"Failed to delete transaction ID {rid} from DB: {e}", exc_info=True)
            QMessageBox.critical(self, "Delete Error", f"An error occurred during final deletion: {e}")
        finally:
            if rid in self.pending_delete_transactions:
                try:
                    del self.pending_delete_transactions[rid]
                except KeyError:
                    pass
            self.refresh_ui()

    # -=- Settings -=-
    def on_date_format_changed(self):
        settings = QSettings("Azralithia", "FinanceTracker")
        current_display_format = settings.value("date_format", "(YYYY-MM-DD) | Year-Month-Day")
        display_format = map_display_format(current_display_format)

        if hasattr(self, "history_page"):
            self.history_page.start_date.setDisplayFormat(display_format)
            self.history_page.end_date.setDisplayFormat(display_format)
            self.history_page.load_page()

        if hasattr(self, "transactions_page"):
            self.transactions_page.date.setDisplayFormat(display_format)
            self.transactions_page.load_recent_transactions() 
        
        if hasattr(self, "show_summary_tab"):
            self.show_summary_tab.start_picker.setDisplayFormat(display_format)
            self.show_summary_tab.end_picker.setDisplayFormat(display_format)
            self.show_summary_tab.refresh_summary()

    def handle_action(self, action_name: str):
        logger.info(f"Action triggered: {action_name}")
        new_title_suffix = ""

        if action_name == "Main":
            self.stack.setCurrentWidget(self.main_page)
            new_title_suffix = ""

        elif action_name == "Settings":
            self.stack.setCurrentWidget(self.settings_page)
            new_title_suffix = " - Settings"

        elif action_name == "Exit":
            self.close()
            return 
        
        elif action_name == "Manage Transactions":
            self.stack.setCurrentWidget(self.transactions_page)
            new_title_suffix = " - Manage Transactions"

        elif action_name == "Show Summary":
            self.stack.setCurrentWidget(self.show_summary_tab)
            new_title_suffix = " - Summary"

        elif action_name == "Transaction Log":
            self.stack.setCurrentWidget(self.history_page)
            new_title_suffix = " - Transaction Log"

        elif action_name == "Export":
            self._open_export_dialog()
            return 
        
        elif action_name == "Import":
            self._open_import_dialog()
            return 
        
        self.setWindowTitle(f"Azralithia Finance Tracker{new_title_suffix}")
    
    def closeEvent(self, event):
        if self.settings.value("load_last_page", False, type=bool): 
            self.settings.setValue("last_page_index", self.stack.currentIndex()) 
        save = self.settings.value("save_filters", False, type=bool)
        if save and hasattr(self, "history_page"):
            filters = {
                "type": self.history_page.type_filter.currentText(),
                "category": self.history_page.category_filter.currentText(),
                "start": self.history_page.start_date.date().toString("yyyy-MM-dd"),
                "end": self.history_page.end_date.date().toString("yyyy-MM-dd"),
                "note": self.history_page.note_search.text()
            }
            self.settings.setValue("history_filters", json.dumps(filters))
        super().closeEvent(event)

    def _load_last_page_viewed(self):
        if self.settings.value("load_last_page", False, type=bool):
            last_page_index = self.settings.value("last_page_index", 0, type=int)
            self.stack.setCurrentIndex(last_page_index)
            page_name_map = {
                self.stack.indexOf(self.main_page): "Main",
                self.stack.indexOf(self.transactions_page): "Manage Transactions",
                self.stack.indexOf(self.show_summary_tab): "Summary",
                self.stack.indexOf(self.history_page): "Transaction Log",
                self.stack.indexOf(self.settings_page): "Settings"
            }
            current_page_widget = self.stack.currentWidget()
            for index, name in page_name_map.items():
                if self.stack.widget(index) == current_page_widget:
                    self.setWindowTitle(f"Azralithia Finance Tracker - {name}")
                    break
    
    def _update_logging_level_from_settings(self, level_str: str):
        level_map = {
            "DEBUG": logging.DEBUG,
            "INFO": logging.INFO,
            "WARNING": logging.WARNING,
            "ERROR": logging.ERROR,
            "CRITICAL": logging.CRITICAL
        }
        new_level = level_map.get(level_str, logging.INFO)
        logger.setLevel(new_level)
        logger.info(f"Logging level initialized/updated to {level_str}")

    # -=- Dialog -=-
    def _open_import_dialog(self):
        dialog = ImportOptionsDialog(db_path="transactions.db", parent=self)
        dialog.exec()

    def _open_export_dialog(self):
        dialog = ExportOptionsDialog(db_path="transactions.db", parent=self)
        dialog.exec()

    # -=- Theme settings -=-
    def toggle_theme(self, light_mode: bool):
        self.settings.setValue("light_mode", bool(light_mode)) 
        theme_mode = "light" if light_mode else "dark"
        
        # Main Stylesheet
        if light_mode:
            self.apply_light_theme()
        else:
            self.apply_dark_theme()
        
        # Propagate theme to pages that support it
        if hasattr(self, 'show_summary_tab'):
            self.show_summary_tab.apply_theme(theme_mode)
        
        if hasattr(self, 'transactions_page'):
            self.transactions_page.apply_theme(theme_mode)

    def apply_dark_theme(self):
        self.setStyleSheet(DARK_MODE)
        self.sidebar.theme_switch._track_dark = True
        self.sidebar.theme_switch.update()

    def apply_light_theme(self):
        self.setStyleSheet(LIGHT_MODE)
        self.sidebar.theme_switch._track_dark = False
        self.sidebar.theme_switch.update()

    def refresh_ui(self):
        self.main_page.load_summary()
        self.show_summary_tab.refresh_summary()
        self.history_page.load_page()
        self.transactions_page.load_recent_transactions()


if __name__ == "__main__":
    app = QApplication(sys.argv)
    app.setStyle("Fusion")  
    window = MainWindow()
    window.show()
    sys.exit(app.exec())